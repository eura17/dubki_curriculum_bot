# -*- coding: utf-8 -*-

import os
from threading import Thread, Lock
from time import sleep
import psycopg2 as sql
import openpyxl as xcl
from abstractions import now
from curriculums import BusesCurriculum, SuburbansCurriculum


class Config:
    """ Класс, отвечающий за настройку бота при первом запуску и последующие
        обновления расписания
    """
    def __init__(self):
        """ При инициализации запускается первый запуск: установка расписания
            автобусов и электричек, запуск потока обновления расписания
            каждые 6 часов и подключение к базе данных
        """
        print('Первый запуск...')
        self.buses_curriculum = BusesCurriculum()
        print('Расписание автобусов загружено...')
        self.suburbans_curriculum = SuburbansCurriculum()
        print('Расписание электричек загружено...')
        self.last_update_time = now()
        self.update_curriculum_thread = Thread(target=self.update_curriculum,
                                               name='update_curriculum_thread',
                                               daemon=True)
        self.update_curriculum_thread.start()
        print('Поток обновления расписания запущен...')
        self.database = DataBase()
        print('База данных подключена...')
        print('Первый запуск прошел успешно. Бот готов к работе!')

    def update_curriculum(self):
        """ Функция для потока обновления расписания каждые 6 часов
        """
        with Lock():
            while True:
                sleep(3600 * 6)
                if now().hour in (0, 1, 2):
                    sleep(3600 * 3)
                self.buses_curriculum = BusesCurriculum()
                self.suburbans_curriculum = SuburbansCurriculum()
                self.last_update_time = now()
                print(f'Расписание обновилось. ({now().strftime("%H:%M:%S")})')


class DataBase:
    """ Класс для работы с базой данных, в которой хранится вся статистика
        по использовнию бота
    """
    def __init__(self):
        """ Подключается к базе данных на Heroku по ссылке и создает
            подключение и курсор
        """
        self.connection_to_db = sql.connect(os.getenv('DATABASE_URL'),
                                            sslmode='require')
        self.cursor_to_db = self.connection_to_db.cursor()
        if int(os.getenv('CREATE_DB')):
            self.create_db()

    def create_db(self):
        """ Создает базу данных с таблицами full_statistics, где хранится
            статистика по использованию команд бота по датам, и
            users_statistics, где хранится статистика по использованию команд
            бота по конкретным пользователям
        """
        self.create_full_statistics_table()
        self.create_users_statistics_table()

    def create_full_statistics_table(self):
        """ Создает таблицу full_statistics в базе данных sql с датой,
            статистикой по вызовам каждой команды и общим количеством вызовов
            бота за день
        """
        self.cursor_to_db.execute('''CREATE TABLE FULL_STATISTICS
                                     (date TEXT PRIMARY KEY,
                                     start_calls INT,
                                     help_calls INT,
                                     buses_calls INT,
                                     slavyanki_calls INT,
                                     trains_calls INT,
                                     file_calls INT,
                                     total_calls INT);''')
        self.connection_to_db.commit()

    def create_users_statistics_table(self):
        """ Создает таблицу users_statistics в базе данных sql с telegram_id
            пользователя, датой начала использования бота, датой последнего
            вызова бота и общим количеством вызовов бота
        """
        self.cursor_to_db.execute('''CREATE TABLE USERS_STATISTICS
                                     (id INT PRIMARY KEY NOT NULL,
                                     start_date TEXT,
                                     lastcall_date TEXT,
                                     total_calls INT);''')
        self.connection_to_db.commit()

    def add_user_data(self, command: str, user_id: int):
        """ Принимает на вход название команды бота command: str и telegram-id
            пользователя user_id: int и добавляет данные в базу данных со
            статистикой по использованию
        """
        today = now().strftime('%d.%m.%Y')
        self.update_full_statistics(today, command)
        self.update_users_statistics(today, user_id)

    def update_full_statistics(self, today: str, command: str):
        """ Принимает на вход сегодняшнюю дату today: str и вызванную команду
            бота command: str и добавляет данные в таблицу full_statistics
        """
        if not self.is_current_date_exists_in_db(today):
            self.create_new_date_in_db(today)
        used_times_cmd, used_times_tc = self.get_used_times_by_date(today,
                                                                    command)
        update_date_cmd = f'''UPDATE FULL_STATISTICS 
                              set {command}_calls = {used_times_cmd+1} 
                              where date LIKE \'{today}\';
                              UPDATE FULL_STATISTICS 
                              set total_calls = {used_times_tc} 
                              where date LIKE \'{today}\';'''
        self.cursor_to_db.execute(update_date_cmd)
        self.connection_to_db.commit()

    def is_current_date_exists_in_db(self, today: str):
        """ Принимает на вход сегодняшнюю дату today: str и проверяет, есть ли
            она в базе данных
        """
        check_date_cmd = f'SELECT EXISTS(SELECT * from FULL_STATISTICS WHERE' \
                         f' date LIKE \'{today}\');'
        self.cursor_to_db.execute(check_date_cmd)
        answer_from_db = self.cursor_to_db.fetchall()
        return answer_from_db[0][0]

    def create_new_date_in_db(self, today: str):
        """ Принимает на вход сегодняшнюю дату today: str и создает строку с
            новой датой в базе данных
        """
        insert_date_cmd = f'''INSERT INTO FULL_STATISTICS 
                             (date, 
                             start_calls, 
                             help_calls, 
                             buses_calls, 
                             slavyanki_calls, 
                             trains_calls, 
                             file_calls, 
                             total_calls) 
                             VALUES 
                             (\'{today}\', 0, 0, 0, 0, 0, 0, 0);'''
        self.cursor_to_db.execute(insert_date_cmd)
        self.connection_to_db.commit()

    def get_used_times_by_date(self, today: str, command: str):
        """ Принимает на вход сегодняшнюю дату today: str и название команды
            бота command: str и возвращает значение, сколько раз эта команда
            была вызвана за сегодня
        """
        commands = {'start': 1,
                    'help': 2,
                    'buses': 3,
                    'slavyanki': 4,
                    'trains': 5,
                    'file': 6,
                    'total': 7}
        get_rows_by_date_cmd = f'SELECT * from FULL_STATISTICS ' \
                               f'WHERE date LIKE \'{today}\';'
        self.cursor_to_db.execute(get_rows_by_date_cmd)
        rows = self.cursor_to_db.fetchall()
        used_times_cmd = rows[0][commands[command]]
        used_times_tc = rows[0][-1]
        return used_times_cmd, used_times_tc

    def update_users_statistics(self, today: str, user_id: int):
        """ Принимает на вход сегодняшнюю дату today: str и telegram-id
            пользователя, вызвавшего бота, user_id: int и добавляет данные
            в таблицу users_statistics
        """
        if not self.is_current_user_exists_in_db(user_id):
            self.create_new_user_in_db(today, user_id)
        amount_of_used_times = self.get_used_times_by_id(user_id)
        update_id_cmd = f'''UPDATE USERS_STATISTICS 
                            set lastcall_date = \'{today}\' 
                            WHERE id = {user_id};
                            UPDATE USERS_STATISTICS 
                            set total_calls = {amount_of_used_times+1} 
                            WHERE id = {user_id};'''
        self.cursor_to_db.execute(update_id_cmd)
        self.connection_to_db.commit()

    def is_current_user_exists_in_db(self, user_id: int):
        """ Принимает на вход telegram-id пользователя user_id: int и
            проверяет, есть ли он в базе данных
        """
        check_id_cmd = f'SELECT EXISTS(SELECT * from USERS_STATISTICS ' \
                       f'WHERE id = {user_id})'
        self.cursor_to_db.execute(check_id_cmd)
        answer_from_db = self.cursor_to_db.fetchall()
        return answer_from_db[0][0]

    def create_new_user_in_db(self, today: str, user_id: int):
        """ Принимает на вход сегодняшнюю дату today: str и telegram-id
            пользователя user_id: int и создает строку с новым пользователем
            в базе данных
        """
        insert_user_cmd = f'''INSERT INTO USERS_STATISTICS 
                              (id, start_date, lastcall_date, total_calls) 
                              VALUES ({user_id}, \'{today}\', \'{today}\', 0);
                           '''
        self.cursor_to_db.execute(insert_user_cmd)
        self.connection_to_db.commit()

    def get_used_times_by_id(self, user_id: int):
        """ Принимает на вход telegram-id пользователя user_id: int и
            возвращает значение, сколько раз пользователь вызывал бота
        """
        get_rows_by_id_cmd = f'SELECT * from USERS_STATISTICS ' \
                             f'WHERE id = {user_id}'
        self.cursor_to_db.execute(get_rows_by_id_cmd)
        rows = self.cursor_to_db.fetchall()
        amount_of_used_times = rows[0][3]
        return amount_of_used_times

    def get_statistics_in_xcl(self):
        """ Выгружает информацию из базы данных и сохраняет ее в виде
            excel-таблицы в файл statistics.xlsx
        """
        wb = xcl.workbook.Workbook()
        wb.active.title = 'full_statistics'
        wb.create_sheet('users_statistics', 1)
        full_statistics_sheet = wb['full_statistics']
        cells = {'A1': 'date',
                 'B1': 'start_calls',
                 'C1': 'help_calls',
                 'D1': 'buses_calls',
                 'E1': 'slavyanki_calls',
                 'F1': 'trains_calls',
                 'G1': 'file_calls',
                 'H1': 'total_calls'}
        for cell in cells:
            full_statistics_sheet[cell] = cells[cell]
        users_statistics_sheet = wb['users_statistics']
        cells = {'A1': 'id',
                 'B1': 'start_date',
                 'C1': 'lastcall_date',
                 'D1': 'total_calls'}
        for cell in cells:
            users_statistics_sheet[cell] = cells[cell]
        for sheet in full_statistics_sheet, users_statistics_sheet:
            sheet_name = 'full_statistics' if sheet is full_statistics_sheet \
                else 'users_statistics'
            rows = self.get_rows_by_table_name(sheet_name)
            for i in range(len(rows)):
                for j in range(len(rows[i])):
                    sheet.cell(row=i + 2, column=j + 1).value = rows[i][j]
        wb.save('statistics.xlsx')

    def get_rows_by_table_name(self, table_name: str):
        """ Принимает на вход название таблицы table_name: str и возвращает
            все данные из базы данных из этой таблицы
        """
        get_full_info_fs_cmd = f'SELECT * FROM {table_name.upper()};'
        self.cursor_to_db.execute(get_full_info_fs_cmd)
        rows = self.cursor_to_db.fetchall()
        return rows
