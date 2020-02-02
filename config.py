# -*- coding: utf-8 -*-

import fitz
import os
import re
import pandas as pd
import openpyxl as xl
import psycopg2 as sql
import datetime as dt
from urllib.request import urlretrieve as download, urlopen as openurl
from requests import get
from threading import Thread, Lock
from bs4 import BeautifulSoup
from time import sleep
from shutil import rmtree as delete


def today():
    return dt.datetime.today() + dt.timedelta(hours=3)


class Bus:
    def __init__(self,
                 timeOfD: list, weekday: str, direction: str,
                 isSlavyanka: bool = False, isAccurate: bool = True):
        self.timeOfDeparture = dt.datetime(year=today().year,
                                           month=today().month,
                                           day=today().day,
                                           hour=timeOfD[0],
                                           minute=timeOfD[1])
        if not isAccurate:
            self.timeOfDeparture += dt.timedelta(minutes=15)
        if timeOfD[0] in (0, 1, 2):
            self.timeOfDeparture += dt.timedelta(days=1)
        self.timeOfArrival = dt.datetime(year=today().year,
                                         month=today().month,
                                         day=today().day,
                                         hour=timeOfD[0],
                                         minute=timeOfD[1]) + \
            dt.timedelta(minutes=15)
        self.weekday = weekday
        self.direction = direction
        self.isSlavyanka = isSlavyanka
        self.isAccurate = isAccurate
        self.trains = []

    def __str__(self):
        toPrint = 'по расписанию {}:{}' if self.isAccurate \
            else 'по прибытии ~ {}:{}'
        hour, minute = self.timeOfDeparture.hour, self.timeOfDeparture.minute
        hour = str(hour) if len(str(hour)) == 2 else '0' + str(hour)
        minute = str(minute) if len(str(minute)) == 2 else '0' + str(minute)
        toPrint = toPrint.format(hour, minute)
        toPrint = toPrint if not self.isSlavyanka \
                          else toPrint + ' (славянка)'
        return toPrint

    def __eq__(self, other):
        if isinstance(other, Bus):
            if self.timeOfDeparture == other.timeOfDeparture:
                return True

    def __lt__(self, other):
        if isinstance(other, Bus):
            if self.timeOfDeparture < other.timeOfDeparture:
                return True
            else:
                return False
        elif isinstance(other, dt.datetime):
            if self.timeOfDeparture < other:
                return True
            else:
                return False

    def __gt__(self, other):
        if isinstance(other, Bus):
            if self.timeOfDeparture > other.timeOfDeparture:
                return True
            else:
                return False
        elif isinstance(other, dt.datetime):
            if self.timeOfDeparture > other:
                return True
            else:
                return False

    def busToPrint(self, forWhat: str = 'buses', delta=None):
        toPrint = '🚌 *{}:{}*' if self.isAccurate \
            else '🚌 по прибытии ~ *{}:{}*'
        hour, minute = self.timeOfDeparture.hour, self.timeOfDeparture.minute
        hour = str(hour) if len(str(hour)) == 2 else '0' + str(hour)
        minute = str(minute) if len(str(minute)) == 2 else '0' + str(minute)
        toPrint = toPrint.format(hour, minute)
        if forWhat == 'slavyanki':
            return toPrint
        toPrint = toPrint if not self.isSlavyanka \
            else toPrint + ' (*славянка*)'
        toPrint += delta
        if len(self.trains) > 0:
            toPrint += '\nудобные пересадки:'
            for train in self.trains:
                if train.mainTime >= today():
                    toPrint += '\n' + train.trainToPrint(forWhat='buses')
        return toPrint


class BusesCurriculum:
    def __init__(self):
        self.weekCurriculum = None
        self.satCurriculum = None
        self.sunCurriculum = None
        self.linkToPDF = None

    def getPdf(self):
        """ Находит pdf-файл с текущим расписанием из кнопки 'Расписание' в
            группе Дубков ('https://vk.com/dubki') и сохраняет его в файл
            Расписание.pdf
        """
        groupPage = BeautifulSoup(openurl('https://vk.com/dubki'),
                                  features='html.parser')
        for a in groupPage.find_all('a', href=True):
            if 'Расписание' in str(a):
                linkToPDF = a['href']
                download(linkToPDF, filename='Расписание.pdf')
                break
        self.linkToPDF = linkToPDF
        self.pdfToTxt()

    @staticmethod
    def pdfToTxt():
        """ Достает текст из pdf-файла с расписанием (Расписание.pdf),
            сохраняет первую страницу (расписание на неделю)
            в txt-файл mon-fri.txt, вторую страницу (расписание на субботу и
            воскресенье) в txt-файл sat-sun.txt (файлы хранятся в папке txt/)
        """
        os.mkdir(os.getcwd() + '/txt')
        pdfDocument = 'Расписание.pdf'
        doc = fitz.open(pdfDocument)
        for i in range(2):
            page = doc.loadPage(i)
            text = page.getText('text')
            pathToSave = 'txt/mon-fri.txt' if i == 0 else 'txt/sat-sun.txt'
            with open(pathToSave, 'w', encoding='utf8') as f:
                print(text, file=f)

    def createCurriculum(self):
        """ Парсит txt-файлы mon-fri.txt и sat-sun.txt (из папки txt/)
            и сохраняет расписание в списки с "сырым" расписанием, которые
            впоследствии преобразуются в датафреймы pandas с расписание
            на будние дни, субботу и воскресенье
        """
        DtoO = True
        DtoOweekTimes, OtoDweekTimes = [], []
        DtoOsatTimes, OtoDsatTimes = [], []
        DtoOsunTimes, OtoDsunTimes = [], []
        toPutDtoO = {'week': DtoOweekTimes,
                     'sat': DtoOsatTimes,
                     'sun': DtoOsunTimes}
        toPutOtoD = {'week': OtoDweekTimes,
                     'sat': OtoDsatTimes,
                     'sun': OtoDsunTimes}
        weekday = ['week', 'sat', 'sun']
        indexOfWeekday = 0

        for file in 'mon-fri.txt', 'sat-sun.txt':
            if file == 'sat-sun.txt':
                indexOfWeekday += 1
            with open('txt/{}'.format(file), 'r', encoding='utf8') as f:
                listOfLines = []
                lineOf3Pairs = [(1000, 1000) for i in range(3)]
                numOfVacantPlace = 0
                for line in f:
                    predict = False
                    slavyanka = False
                    line = line.strip()
                    if line == 'по прибыт.':
                        predict = True
                        hours, minutes = dubki[0]
                        time = [hours, minutes]
                    elif line == '----':
                        time = [404, 404]
                    elif re.search(pattern=r'\d\d:\d\d',
                                   string=line) is not None or \
                            re.search(pattern=r'\d:\d\d',
                                      string=line) is not None:
                        if line[-2:] == '**':
                            slavyanka = True
                            line = line[:-2]
                        elif line[-1] == '*':
                            slavyanka = True
                            line = line[:-1]
                        line = list(map(int, line.split(':')))
                        hours, minutes = int(line[0]), int(line[1])
                        time = [hours, minutes]
                    else:
                        continue
                    if DtoO:
                        DtoO = False
                        dubki = (time, predict, slavyanka)
                    else:
                        DtoO = True
                        lineOf3Pairs[numOfVacantPlace] = (dubki,
                                                          (time, predict, slavyanka))
                        numOfVacantPlace += 1

                    if numOfVacantPlace == 3:
                        for pair in lineOf3Pairs:
                            if pair[0][0][0] in {0, 1, 2, 3, 4}:
                                pair[0][0][0] += 24
                        lineOf3Pairs.sort()
                        for pair in lineOf3Pairs:
                            if pair[0][0][0] in {24, 25, 26, 27, 28}:
                                pair[0][0][0] -= 24
                        listOfLines.append(lineOf3Pairs)
                        lineOf3Pairs = [(1000, 1000) for i in range(3)]
                        numOfVacantPlace = 0

                valOfPreLineInFirstPos = 0
                for line in listOfLines:
                    firstPosOfLine = line[0][0][0][0]
                    if firstPosOfLine < valOfPreLineInFirstPos:
                        indexOfWeekday += 1
                    valOfPreLineInFirstPos = firstPosOfLine
                    for pair in line:
                        toPutDtoO[weekday[indexOfWeekday]].append(pair[0])
                        toPutOtoD[weekday[indexOfWeekday]].append(pair[1])

        self.weekCurriculum = self.listsTopdDataframe(DtoOweekTimes,
                                                      OtoDweekTimes,
                                                      weekday='week')
        self.satCurriculum = self.listsTopdDataframe(DtoOsatTimes,
                                                     OtoDsatTimes,
                                                     weekday='sat')
        self.sunCurriculum = self.listsTopdDataframe(DtoOsunTimes,
                                                     OtoDsunTimes,
                                                     weekday='sun')

        delete(os.getcwd() + '/txt')

    @staticmethod
    def listsTopdDataframe(timesDubki: list, timesOdintsovo: list,
                           weekday: str):
        """ Превращает списки с "сырым" расписанием в виде списков в
            датафрейм pandas, каждая ячейка которого принадлежит классу Bus
            или имеет значение NaN, при этом таблица имеет отсортированный вид
        """
        times = {'Дубки-Одинцово': [], 'Одинцово-Дубки': []}
        for timetable in timesDubki, timesOdintsovo:
            for line in timetable:
                direction = 'Дубки-Одинцово' if timetable is timesDubki \
                            else 'Одинцово-Дубки'
                if line[0] != [404, 404]:
                    bus = Bus(timeOfD=line[0],
                              weekday=weekday,
                              direction=direction,
                              isAccurate=not line[1],
                              isSlavyanka=line[2])
                    times[direction].append(bus)
            times[direction].sort()
        curriculum = pd.concat(
            [pd.DataFrame({'Дубки-Одинцово': times['Дубки-Одинцово']}),
             pd.DataFrame({'Одинцово-Дубки': times['Одинцово-Дубки']})],
            axis=1)
        return curriculum

    def getBusesForHour(self, direction: str):
        """ Находит все автобусы в пределах ближайшего часа и возвращает
            список из этих автобусов класса Bus
        """
        start = today()
        end = start + dt.timedelta(hours=1)
        relevantCurriculum = self.getRelevantCurriculum()
        relevantBuses = []
        for bus in relevantCurriculum[direction]:
            if not isinstance(bus, Bus):
                continue
            if start < bus.timeOfDeparture < end or \
                    start + dt.timedelta(days=1) < bus.timeOfDeparture < end + dt.timedelta(days=1):
                delta = abs((start.hour * 60 + start.minute) -
                            (bus.timeOfDeparture.hour * 60 + bus.timeOfDeparture.minute))
                if delta > 60:
                    delta = 60 - delta % 60
                textDelta = ' – через {}'
                if delta == 60:
                    addDelta = textDelta.format('1 час')
                elif delta == 0:
                    addDelta = ' – сейчас'
                elif 11 <= delta <= 19:
                    addDelta = textDelta.format('{} минут'.format(delta))
                elif delta % 10 == 1:
                    addDelta = textDelta.format('{} минуту'.format(delta))
                elif delta % 10 in [2, 3, 4]:
                    addDelta = textDelta.format('{} минуты'.format(delta))
                elif delta % 10 in [5, 6, 7, 8, 9, 0]:
                    addDelta = textDelta.format('{} минут'.format(delta))
                relevantBuses.append(bus.busToPrint(delta=addDelta))
        return relevantBuses

    def getSlavyanki(self, direction: str):
        relevantCurriculum = self.getRelevantCurriculum()
        relevantBuses = []
        for bus in relevantCurriculum[direction]:
            if not isinstance(bus, Bus):
                continue
            if bus.isSlavyanka and today() < bus.timeOfDeparture:
                relevantBuses.append(bus.busToPrint(forWhat='slavyanki'))
        return relevantBuses

    def getRelevantCurriculum(self):
        if today().isoweekday() in (1, 2, 3, 4, 5) or today().isoweekday() == 5 \
                and today().hour in (0, 1, 2):
            relevantCurriculum = self.weekCurriculum
        elif today().isoweekday() == 6 or today().isoweekday() == 6 \
                and today().hour in (0, 1, 2):
            relevantCurriculum = self.satCurriculum
        elif today().isoweekday() == 7 or today().isoweekday() == 7 \
                and today().hour in (0, 1, 2):
            relevantCurriculum = self.sunCurriculum
        return relevantCurriculum


class Train:
    def __init__(self,
                 time: list,  direction: str, name: str, type: str,
                 nextDay: bool = False):
        suburbanTypes = {'stdplus': 'Стандарт плюс',
                         'suburban': 'Обычная электричка',
                         'rex': 'Экспресс РЭКС',
                         'mcd1': 'Иволга',
                         'aerstd': 'Аэроэкспресс'}
        self.mainTime = dt.datetime(year=today().year,
                                    month=today().month,
                                    day=today().day,
                                    hour=time[0],
                                    minute=time[1])
        self.direction = direction
        if nextDay:
            self.mainTime += dt.timedelta(days=1)
        self.suburbanName = name
        self.suburbanType = suburbanTypes[type]
        self.stops = {}
        self.buses = {'week': [],
                      'sat': [],
                      'sun': []}

    def __str__(self):
        hour = self.mainTime.hour if len(str(self.mainTime.hour)) == 2 \
            else '0' + str(self.mainTime.hour)
        minute = self.mainTime.minute if len(str(self.mainTime.minute)) == 2 \
            else '0' + str(self.mainTime.minute)
        toPrint = '{} ({}) – {}:{}\n'.format(self.suburbanName,
                                          self.suburbanType,
                                          hour, minute)
        for station in self.stops:
            hour = self.stops[station].hour if len(str(self.stops[station].hour)) == 2 \
                else '0' + str(self.stops[station].hour)
            minute = self.stops[station].minute if len(str(self.stops[station].minute)) == 2 \
                else '0' + str(self.stops[station].minute)
            toPrint += '{} – {}:{} '.format(station, hour, minute)
        return toPrint

    def __eq__(self, other):
        if isinstance(other, Train):
            if self.mainTime == other.mainTime:
                return True
            else:
                return False

    def __lt__(self, other):
        if isinstance(other, Train):
            if self.mainTime < other.mainTime:
                return True
            else:
                return False
        elif isinstance(other, dt.datetime):
            if self.mainTime < other:
                return True
            else:
                return False

    def __gt__(self, other):
        if isinstance(other, Train):
            if self.mainTime > other.mainTime:
                return True
            else:
                return False
        elif isinstance(other, dt.datetime):
            if self.mainTime > other:
                return True
            else:
                return False

    def trainToPrint(self, forWhat: str = 'buses', station: str = None):
        if forWhat == 'buses':
            hour = self.mainTime.hour if len(str(self.mainTime.hour)) == 2 \
                else '0' + str(self.mainTime.hour)
            minute = self.mainTime.minute if len(str(self.mainTime.minute)) == 2 \
                else '0' + str(self.mainTime.minute)
            toPrint = '🚆 *{}:{} – {}:\n*'.format(hour, minute,
                                             self.suburbanName)
            stations = ('Кунцево', 'Фили', 'Беговая', 'Белорусский вокзал') if self.direction == 'Одинцово-Москва' \
                else ('Белорусский вокзал', 'Беговая', 'Фили', 'Кунцево')
            for station in stations:
                if station in self.stops:
                    hour = self.stops[station].hour if len(str(self.stops[station].hour)) == 2 \
                        else '0' + str(self.stops[station].hour)
                    minute = self.stops[station].minute if len(str(self.stops[station].minute)) == 2 \
                        else '0' + str(self.stops[station].minute)
                    toPrint += '*{}* – {}:{}, '.format(station,
                                                       hour,
                                                       minute)
            toPrint = toPrint[:-2]
        elif forWhat == 'trains':
            toPrint = '🚆 *{}* ({})'.format(self.suburbanName,
                                            self.suburbanType)
            toPrint += ':\nотправление в *{}:{}*, прибытие в *{}:{}*'
            departureHour = self.mainTime.hour if len(str(self.mainTime.hour)) == 2 \
                else '0' + str(self.mainTime.hour)
            departureMinute = self.mainTime.minute if len(str(self.mainTime.minute)) == 2 \
                else '0' + str(self.mainTime.minute)
            arrivalHour = self.stops[station].hour if len(str(self.stops[station].hour)) == 2 \
                else '0' + str(self.stops[station].hour)
            arrivalMinute = self.stops[station].minute if len(str(self.stops[station].minute)) == 2 \
                else '0' + str(self.stops[station].minute)
            if self.direction == 'Москва-Одинцово':
                departureHour, arrivalHour = arrivalHour, departureHour
                departureMinute, arrivalMinute = arrivalMinute, departureMinute
            toPrint = toPrint.format(departureHour, departureMinute,
                                     arrivalHour, arrivalMinute)

        return toPrint


class TrainsCurriculum:
    def __init__(self):
        self.suburbanCurriculum = None
        self.trains = None

    def makeRequest(self, stationFrom: str, stationTo: str,
                    direction: str, nextDay: bool = False):
        stations = {'Одинцово': 'c10743',
                    'Кунцево': 's9601728',
                    'Фили': 's9600821',
                    'Беговая': 's9601666',
                    'Белорусский вокзал': 's2000006'}
        apikey = os.getenv('APIKEY_YANDEX')
        url = 'https://api.rasp.yandex.net/v3.0/search/?apikey={}&from={}&to={}&date={}'
        default = '&format=json&limit=1000&lang=ru_RU&system=yandex&show_systems=yandex'
        now = today()
        if nextDay:
            now += dt.timedelta(days=1)
        relevantTrains = []

        response = get(url.format(apikey,
                                  stations[stationFrom],
                                  stations[stationTo],
                                  now.date())
                       + default).json()
        if 'error' not in response:
            trains = response['segments']
            amountOfTrains = 0
            for train in trains:
                amountOfTrains += 1
                begin, end = train['departure'].find('T') + 1, \
                             train['departure'].find('T') + 6
                departureTime = train['departure'][begin:end]
                departureTime = list(map(int, departureTime.split(':')))
                begin, end = train['arrival'].find('T') + 1, \
                             train['arrival'].find('T') + 6
                arrivalTime = train['arrival'][begin:end]
                arrivalTime = list(map(int, arrivalTime.split(':')))
                if direction == 'Одинцово-Москва':
                    time = departureTime
                    stopTime = arrivalTime
                    stop = stationTo
                elif direction == 'Москва-Одинцово':
                    time = arrivalTime
                    stopTime = departureTime
                    stop = stationFrom
                if (time[0] in {0, 1, 2} or stopTime[0] in {0, 1, 2}) and \
                        amountOfTrains >= 30:
                    continue

                mainTime = dt.datetime(year=now.year,
                                       month=now.month,
                                       day=now.day,
                                       hour=time[0],
                                       minute=time[1])
                suburbanName = train['thread']['title']
                if suburbanName == 'Москва (Белорусский вокзал) — Одинцово':
                    suburbanName = 'Лобня — Одинцово'
                suburbanType = train['thread']['transport_subtype']['code']

                found = False
                for tr in self.trains:
                    if tr.suburbanName == suburbanName and \
                       tr.mainTime == mainTime:
                        tr.stops[stop] = dt.datetime(year=now.year,
                                                     month=now.month,
                                                     day=now.day,
                                                     hour=stopTime[0],
                                                     minute=stopTime[1])
                        found = True
                        break
                if not found:
                    currentTrain = Train(time, direction,
                                         suburbanName, suburbanType, nextDay=nextDay)
                    currentTrain.stops[stop] = dt.datetime(year=now.year,
                                                           month=now.month,
                                                           day=now.day,
                                                           hour=stopTime[0],
                                                           minute=stopTime[1])
                    relevantTrains.append(currentTrain)
        return relevantTrains

    def createCurriculum(self):
        df = pd.DataFrame({'nothing': [1, 2, 3]})
        for departure in 'Одинцово', 'Москва':
            self.trains = []
            direction = '{}-{}'.format(departure,
                                           'Москва' if departure == 'Одинцово'
                                            else 'Одинцово')
            for station in ['Кунцево', 'Фили', 'Беговая', 'Белорусский вокзал']:
                start = station if departure == 'Москва' else 'Одинцово'
                stop = station if departure == 'Одинцово' else 'Одинцово'
                self.trains += self.makeRequest(stationFrom=start,
                                                stationTo=stop,
                                                direction=direction)
                self.trains += self.makeRequest(stationFrom=start,
                                                stationTo=stop,
                                                direction=direction,
                                                nextDay=True)
            self.trains.sort()
            df = pd.concat([df, pd.DataFrame({direction: self.trains})],
                           axis=1,
                           sort=True)
        self.suburbanCurriculum = df.drop(['nothing'], axis=1)

    def getTrainsForHour(self, direction: str):
        start = today()
        end = today() + dt.timedelta(hours=1)
        station = direction.replace('Одинцово', '').replace('-', '')
        direction = 'Одинцово-Москва' if direction in \
                                         ('Одинцово-Кунцево', 'Одинцово-Фили', 'Одинцово-Беговая', 'Одинцово-Белорусский вокзал') \
            else 'Москва-Одинцово'
        relevantTrains = []
        for train in self.suburbanCurriculum[direction]:
            if not isinstance(train, Train):
                continue
            if station in train.stops:
                trainTime = train.mainTime if direction == 'Одинцово-Москва' else train.stops[station]
                if start < trainTime < end or \
                        start + dt.timedelta(days=1) < trainTime < end + dt.timedelta(days=1) \
                        and start.hour in (23, 0, 1, 2):
                    relevantTrains.append(train.trainToPrint(forWhat='trains',
                                                             station=station))
        return relevantTrains


class Config:
    def __init__(self):
        self.busesCurriculum = BusesCurriculum()
        self.trainsCurriculum = TrainsCurriculum()
        self.lastUpdate = today()

    @staticmethod
    def createDataBase():
        """ Подключается к базе данных Postgresql (на Heroku), работает только
        в том случае, когда переменная окружения имеет значение не 0
        """
        con = sql.connect(os.getenv('DATABASE_URL'),
                          sslmode='require')
        cur = con.cursor()
        # создание таблицы full_statistics
        cur.execute('''CREATE TABLE FULL_STATISTICS
                    (date TEXT PRIMARY KEY,
                    start_calls INT,
                    help_calls INT,
                    buses_calls INT,
                    slavyanki_calls INT,
                    trains_calls INT,
                    file_calls INT,
                    total_calls INT);''')
        # создание таблицы users_statistics
        cur.execute('''CREATE TABLE USERS_STATISTICS
                    (id INT PRIMARY KEY NOT NULL,
                    start_date TEXT,
                    lastcall_date TEXT,
                    total_calls INT);''')
        con.commit()
        con.close()

    def linkBusesAndTrains(self):
        for daysCurriculum in [self.busesCurriculum.weekCurriculum,
                               self.busesCurriculum.satCurriculum,
                               self.busesCurriculum.sunCurriculum]:
            if daysCurriculum is self.busesCurriculum.sunCurriculum:
                a = 1
            for directionBuses in daysCurriculum:
                for bus in daysCurriculum[directionBuses]:
                    if not isinstance(bus, Bus) or bus.isSlavyanka:
                        continue
                    directionTrains = 'Одинцово-Москва' if directionBuses == 'Дубки-Одинцово' \
                                      else 'Москва-Одинцово'
                    trainsForBus = []
                    for train in self.trainsCurriculum.suburbanCurriculum[directionTrains]:
                        if not isinstance(train, Train):
                            continue
                        delta = train.mainTime - bus.timeOfArrival if directionBuses == 'Дубки-Одинцово' \
                                else bus.timeOfDeparture - train.mainTime
                        deltaHours = delta.seconds // 3600
                        if deltaHours == 0 and delta.days == 0 and not delta.seconds == 0:
                            trainsForBus.append(train)
                        elif delta.days == 0 and delta.seconds == 0 and directionBuses == 'Дубки-Одинцово':
                            trainsForBus.append(train)
                    if len(trainsForBus) > 3:
                        if directionTrains == 'Москва-Одинцово':
                            bus.trains = trainsForBus[-1:-4:-1]
                            bus.trains.reverse()
                        elif directionTrains == 'Одинцово-Москва':
                            bus.trains = trainsForBus[:3]
                    else:
                        bus.trains = trainsForBus
                    for train in bus.trains:
                        if len(train.buses[bus.weekday]) < 3:
                            train.buses[bus.weekday].append(bus)

    def firstSetup(self):
        """ Основная функция для настройки при первоначальном деплое на Heroku.
            Засекает время, затраченное на подготовку расписаний, и
            запускает два побочных daemon-потока, отвечающих за постоянное обновление
            для электричек - каждый час, для автобусов - каждые 12 часов
        """
        print('Настройка началась...')
        self.busesCurriculum.getPdf()
        print('Pdf-файл с расписанием автобусов загружен...')
        self.busesCurriculum.createCurriculum()
        print('Расписание автобусов создано и сохранено...')
        self.trainsCurriculum.createCurriculum()
        print('Расписание электричек загружено и сохранено...')
        self.linkBusesAndTrains()
        print('Привязка электричек к автобусам прошла успешно...')
        self.lastUpdate = today()

        def updateCurriculum():
            with Lock():
                while True:
                    sleep(3600 * 6)
                    if today().hour in (0, 1, 2):
                        sleep(3600 * 3)
                    self.busesCurriculum.getPdf()
                    self.busesCurriculum.createCurriculum()
                    self.trainsCurriculum.createCurriculum()
                    self.linkBusesAndTrains()
                    self.lastUpdate = today()

        updateCurriculumThread = Thread(target=updateCurriculum,
                                        name='updateCurriculum',
                                        daemon=True)
        updateCurriculumThread.start()
        print('Поток обновления расписания запущен...')

        if bool(int((os.getenv('CREATE_DB')))):
            self.createDataBase()
            print('База данных создана...')
        else:
            try:
                con = sql.connect(os.getenv('DATABASE_URL'),
                                  sslmode='require')
                con.close()
                print('База данных подключена...')
            except Exception as e:
                print('Проблемы с подключением к базе данных...')
                print(e)

        print('Первая настройка прошла успешно. Бот готов к работе.')


class Answers:
    def __init__(self):
        self.config = Config()
        self.config.firstSetup()

    def startAnswer(self):
        answer = '''Привет, дубчанин!
Я бот, с помощью которого ты сможешь узнать расписание ближайших автобусов и электричек.
Да, тебе не придется лезть в группу в ВК, я сделаю это за тебя.
Инструкция по работе со мной /help

P.S. По всем вопросам и предложениям писать @eura71'''

    def helpAnswer(self):
        answer = '''Доступные команды:
/buses - список автобусов на ближайший час и электричек, подходящих для пересадки
/slavyanki - список славянок на сегодня
/trains - список электричек на ближайший час
/file - PDF-файл с расписанием'''
        return answer

    def busesAnswer(self, direction: str):
        buses = self.config.busesCurriculum.getBusesForHour(direction)
        if len(buses) == 0:
            answer = 'Автобусов в ближайший час нет :('
        else:
            answer = 'Автобусы от Дубков до Одинцово:\n' if direction=='Дубки-Одинцово' \
                else 'Автобусы от Одинцово до Дубков:\n'
            for bus in buses:
                answer += '\n{}\n'.format(bus)
            answer += '\nВсе электрички /trains'
        return answer

    def slavyankiAnswer(self, direction):
        slavyanki = self.config.busesCurriculum.getSlavyanki(direction)
        if len(slavyanki) == 0:
            answer = 'Славянок на сегодня больше нет :('
        else:
            answer = 'Автобусы от Дубков до Славянского бульвара' if direction == 'Дубки-Одинцово' \
                else 'Автобусы от Славянского бульвара до Дубков'
            for slavyanka in slavyanki:
                answer += '\n{}\n'.format(slavyanka)
        return answer

    def trainsAnswer(self, direction):
        trains = self.config.trainsCurriculum.getTrainsForHour(direction)
        if len(trains) == 0:
            answer = 'Электричек в ближайший час нет :('
        else:
            grammarForStations = {'Одинцово-Кунцево': 'от Одинцово до Кунцево',
                                  'Одинцово-Фили': 'от Одинцово до Филей',
                                  'Одинцово-Беговая': 'от Одинцово до Беговой',
                                  'Одинцово-Белорусский вокзал': 'от Одинцово до Белорусского вокзала',
                                  'Белорусский вокзал-Одинцово': 'от Белорусского возкала до Одинцово',
                                  'Беговая-Одинцово': 'от Беговой до Одинцово',
                                  'Фили-Одинцово': 'от Филей до Одинцово',
                                  'Кунцево-Одинцово': 'от Кунцево до Одинцово'}
            answer = 'Электрички {}:\n'.format(grammarForStations[direction])
            for train in trains:
                answer += '\n{}\n'.format(train)
            answer += '\nБлижайшие автобусы /buses'
        return answer


class Admin:
    def __init__(self, myID=281452837):
        self.id = myID

    def checkUpdates(self, answers: Answers):
        """ Ответ на специальную скрытую команду, доступную только админу,
            /check_updates - выдает время последнего обновления расписания
            электричек и автобусов
        """
        answer = 'Последнее обновление расписания: *{}:{} {}*'.\
            format(answers.config.lastUpdate.hour,
                   answers.config.lastUpdate.minute,
                   answers.config.lastUpdate.strftime('%d.%m.%Y'))
        return answer

    def userData(self, userID: int, func: str):
        """ Добавляет статистику использования бота в базу данных,
            для этого использует таблицу FULL_STATISTICS для сохранения
            информации по общему использованию всех команд бота и
            таблицу USERS_STATISTICS для сохранения информации по каждому пользователю
        """
        now = today().strftime('%d.%m.%Y')
        con = sql.connect(os.getenv('DATABASE_URL'),
                          sslmode='require')
        cur = con.cursor()
        # работа с таблицей full_statistics
        funcs = {'start': 1,
                 'help': 2,
                 'buses': 3,
                 'slavyanki': 4,
                 'trains': 5,
                 'file': 6,
                 'total': 7}
        cur.execute('SELECT * from FULL_STATISTICS WHERE date LIKE \'{}\';'.
                    format(now))
        rows = cur.fetchall()
        if len(rows) == 0:
            cur.execute('INSERT INTO FULL_STATISTICS (date, start_calls, help_calls, buses_calls, slavyanki_calls, trains_calls, file_calls, total_calls) VALUES (\'{}\', {}, {}, {}, {}, {}, {}, {});'.
                format(now, 0, 0, 0, 0, 0, 0, 0))
            con.commit()
            cur.execute('SELECT * from FULL_STATISTICS WHERE date LIKE \'{}\';'.
                format(now))
            rows = cur.fetchall()
        cur.execute('UPDATE FULL_STATISTICS set {}_calls = {} where date LIKE \'{}\';'.
            format(func, int(rows[0][funcs[func]]) + 1, now))
        cur.execute('UPDATE FULL_STATISTICS set total_calls = {} where date LIKE \'{}\';'.
            format(rows[0][funcs['total']] + 1, now))
        con.commit()
        # работа с таблицей users_statistics
        cur.execute('SELECT * from USERS_STATISTICS WHERE id = {}'.
                    format(userID))
        rows = cur.fetchall()
        if len(rows) == 0:
            cur.execute('INSERT INTO USERS_STATISTICS (id, start_date, total_calls) VALUES ({}, \'{}\', {});'.
                format(userID, now, 0))
            con.commit()
            cur.execute('SELECT * from USERS_STATISTICS WHERE id = {}'.
                        format(userID))
            rows = cur.fetchall()
        cur.execute('UPDATE USERS_STATISTICS set lastcall_date = \'{}\' WHERE id = {};'.
            format(now, userID))
        cur.execute('UPDATE USERS_STATISTICS set total_calls = {} WHERE id = {};'.
                    format(int(rows[0][3]) + 1, userID))
        con.commit()
        con.close()

    @staticmethod
    def createStatisticsXcl():
        """ Выгружает информацию из базы данных Postgresql (на Heroku) и
        сохраняет ее в файл statistics.xlsx для того, чтобы бот мог ответить
        на специальную команду, доступную только админу, /get_statistics
        """
        wb = xl.workbook.Workbook()
        wb.active.title = 'full_statistics'
        wb.create_sheet('users_statistics', 1)
        con = sql.connect(os.getenv('DATABASE_URL'),
                          sslmode='require')
        cur = con.cursor()
        # работа с full_statistics
        fullStatisticsSheet = wb['full_statistics']
        names = {'A1': 'date',
                 'B1': 'start_calls',
                 'C1': 'help_calls',
                 'D1': 'buses_calls',
                 'E1': 'slavyanki_calls',
                 'F1': 'trains_calls',
                 'G1': 'file_calls',
                 'H1': 'total_calls'}
        for elem in names:
            fullStatisticsSheet[elem] = names[elem]
        cur.execute('SELECT * FROM FULL_STATISTICS;')
        rows = cur.fetchall()
        for i in range(len(rows)):
            for j in range(len(rows[i])):
                fullStatisticsSheet.cell(row=i + 2, column=j + 1).value = rows[i][j]
        # работа с users_statistics
        usersStatisticsSheet = wb['users_statistics']
        names = {'A1': 'id',
                 'B1': 'start_date',
                 'C1': 'lastcall_date',
                 'D1': 'total_calls'}
        for elem in names:
            usersStatisticsSheet[elem] = names[elem]
        cur.execute('SELECT * FROM USERS_STATISTICS;')
        rows = cur.fetchall()
        for i in range(len(rows)):
            for j in range(len(rows[i])):
                usersStatisticsSheet.cell(row=i + 2, column=j + 1).value = rows[i][j]
        con.close()
        wb.save('statistics.xlsx')
